from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


class ImportFileError(Exception):
    """Raised when an import file cannot be parsed."""


class ExcelParser:
    """
    Generic XLSX parser.

    This parser knows nothing about Answers, Datapoints,
    Framework Nodes, Stakeholders, or Emission Factors.

    The parser accepts either a filesystem path or a file-like
    object. File-like objects allow Django storage backends such
    as S3/object storage to be used without requiring a physical
    filesystem path.
    """

    SUPPORTED_EXTENSIONS = {".xlsx"}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB

    @classmethod
    def _validate_file_size(cls, file_source):
        size = getattr(file_source, "size", None)

        if size is not None and size > cls.MAX_FILE_SIZE:
            raise ImportFileError(
                "The uploaded file is too large. "
                "Maximum allowed size is 10 MB."
            )

    def parse(self, file_source):
        """
        Parse an XLSX file from a filesystem path or file-like object.

        File-like objects are used by Django storage backends so that
        parsing does not depend on a physical filesystem path.

        Parsed rows are yielded incrementally so callers can process
        large files in bounded chunks instead of storing all rows
        in memory at once.
        """

        self._validate_file_size(file_source)

        if isinstance(file_source, (str, Path)):
            path = Path(file_source)

            self._validate_file(path)

            workbook_source = path
        else:
            self._validate_file_name(file_source)
            workbook_source = file_source

        try:
            workbook = load_workbook(
                filename=workbook_source,
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ImportFileError(
                "The uploaded Excel file could not be read."
            ) from exc

        try:
            worksheet = workbook.active

            row_iterator = worksheet.iter_rows(
                values_only=True,
            )

            try:
                header_row = next(row_iterator)
            except StopIteration:
                raise ImportFileError(
                    "The Excel file does not contain any rows."
                )

            headers = self._get_headers(header_row)

            for excel_row_number, values in enumerate(
                row_iterator,
                start=2,
            ):
                if self._is_blank_row(values):
                    continue

                raw_data = {}

                for index, header in enumerate(headers):
                    value = (
                        values[index]
                        if index < len(values)
                        else None
                    )

                    raw_data[header] = self._json_safe(value)

                yield {
                    "row_number": excel_row_number,
                    "raw_data": raw_data,
                }

        finally:
            workbook.close()

    @classmethod
    def _validate_file(cls, path):
        if not path.exists():
            raise ImportFileError(
                "The uploaded file could not be found."
            )

        if path.suffix.lower() not in cls.SUPPORTED_EXTENSIONS:
            raise ImportFileError(
                "Unsupported file type. Only .xlsx files are supported."
            )

        if path.stat().st_size > cls.MAX_FILE_SIZE:
            raise ImportFileError(
                "The uploaded file is too large. "
                "Maximum allowed size is 10 MB."
            )

    @classmethod
    def _validate_file_name(cls, file_source):
        file_name = getattr(file_source, "name", None)

        if not file_name:
            raise ImportFileError(
                "The uploaded file must have a .xlsx extension."
            )

        if Path(file_name).suffix.lower() not in cls.SUPPORTED_EXTENSIONS:
            raise ImportFileError(
                "Unsupported file type. Only .xlsx files are supported."
            )

    @staticmethod
    def _get_headers(header_row):
        headers = []

        for value in header_row:
            if value is None:
                headers.append("")
            else:
                headers.append(str(value).strip())

        while headers and headers[-1] == "":
            headers.pop()

        if not headers:
            raise ImportFileError(
                "The Excel file does not contain a header row."
            )

        if any(not header for header in headers):
            raise ImportFileError(
                "The Excel header row contains an empty column name."
            )

        if len(headers) != len(set(headers)):
            raise ImportFileError(
                "The Excel header row contains duplicate column names."
            )

        return headers

    @staticmethod
    def _is_blank_row(values):
        return all(
            value is None or str(value).strip() == ""
            for value in values
        )

    @staticmethod
    def _json_safe(value):
        if isinstance(value, datetime):
            if (
                value.hour == 0
                and value.minute == 0
                and value.second == 0
                and value.microsecond == 0
            ):
                return value.date().isoformat()

            return value.isoformat()

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, Decimal):
            return str(value)

        if isinstance(value, bytes):
            return value.decode(
                "utf-8",
                errors="replace",
            )

        return value